import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np


workbook = load_workbook(r'C:\Users\user\Desktop\23-24 Sem1\STAT3609\asm 3 Q8.xlsx')#load your file name
worksheet = workbook["worksheet1"]# load your sheet

#this is the input
expected_return=[worksheet[f'B{i}'].value for i in range(1,9)]
Beta=[worksheet[f'C{i}'].value for i in range(1,9)]
Residual_variance=[worksheet[f'D{i}'].value for i in range(1,9)]
Risk_free=worksheet['G1'].value
Market_variance=worksheet['G2'].value

A = Market_variance * sum((Beta[i] * (expected_return[i] - Risk_free) / Residual_variance[i]) for i in range(len(expected_return)))/(1+Market_variance*sum(Beta[i] ** 2/Residual_variance[i] for i in range(len(expected_return))))
short_selling=input("Y/N")

if short_selling=="Y":
    weight=[beta/non_systematic_risk*((ret-Risk_free)/beta  - A ) for ret,beta,non_systematic_risk in zip(expected_return,Beta,Residual_variance)]
    print(weight)
    weight=[i/sum(weight)for i in weight]
    print(weight)

else:
    z=[]
    for i in range(len(expected_return)):
        R=(expected_return[i]-Risk_free)/Beta[i]
        if (R > A):
            z.append((expected_return[i]-Risk_free)/Beta[i])
        else:
            z.append(0)
    print(z)
    weight = []

    for ret, beta, non_systematic_risk in zip(z, Beta, Residual_variance):
        if ret == 0:
            weight.append(0)
        else:
            weight.append(beta / non_systematic_risk * (ret - A))
    print(weight)
    weight=[i/sum(weight)for i in weight]

        
print(weight)


        

        
